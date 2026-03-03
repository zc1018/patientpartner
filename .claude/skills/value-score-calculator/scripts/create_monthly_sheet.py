#!/usr/bin/env python3
"""
创建价值分月度sheet
根据系统价值分统计数据，自动生成月度进度sheet
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import numbers
import sys
import os

# 添加脚本目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from calculate_progress import calculate_progress, calculate_total_progress, SYSTEM_MAPPING


def create_monthly_sheet(
    value_file: str,
    stats_file: str,
    month_name: str,
    stats_date: str = None
):
    """
    创建新的月度sheet

    Args:
        value_file: 价值分通晒表格.xlsx 路径
        stats_file: 系统价值分统计.xlsx 路径
        month_name: 月份名称（如 "2月"、"3月"）
        stats_date: 统计表中的日期筛选（如 "2026-03-02"），None表示使用最新数据
    """
    print(f"=== 创建 {month_name} 价值分进度表 ===")
    print()

    # 读取数据
    xls_value = pd.ExcelFile(value_file)
    df_init = pd.read_excel(xls_value, sheet_name='初始值')
    df_jan = pd.read_excel(xls_value, sheet_name='1月')  # 用1月作为模板获取车厢、产品信息
    df_stats = pd.read_excel(stats_file, sheet_name='全部周期')

    # 如果指定了日期，筛选数据
    if stats_date:
        df_stats = df_stats[df_stats['日期'] == stats_date]
        if df_stats.empty:
            print(f"错误: 未找到日期 {stats_date} 的数据")
            return

    # 准备数据
    month_data = []

    for idx, row in df_jan.iterrows():
        system_name = row['系统']
        carriage = row['车厢']
        product = row['产品']

        print(f"处理系统: {system_name}")

        # 从初始值获取现状和目标
        init_row = df_init[df_init['系统'] == system_name].iloc[0]

        l1_current = init_row['L1 现状']
        l1_target = init_row['L1 目标']
        l2_current = init_row['L2 现状']
        l2_target = init_row['L2 目标']
        l3_current = init_row['L3 现状']
        l3_target = init_row['L3 目标']

        # 从统计表获取当月数据
        mapped_name = SYSTEM_MAPPING.get(system_name)
        system_stats = df_stats[df_stats['系统名称'] == mapped_name]

        if not system_stats.empty:
            l1_month = float(system_stats.iloc[0]['L1 商业价值分'])
            l2_month = float(system_stats.iloc[0]['L2 渗透率'])
            l3_month = float(system_stats.iloc[0]['L3 系统分'])
            print(f"  ✓ 找到数据: L1={l1_month}, L2={l2_month}, L3={l3_month}")
        else:
            print(f"  ⚠ 未找到 {mapped_name} 的统计数据")
            l1_month = np.nan
            l2_month = np.nan
            l3_month = np.nan

        # 计算进度
        l1_progress = calculate_progress(l1_current, l1_month, l1_target)
        l2_progress = calculate_progress(l2_current, l2_month, l2_target)
        l3_progress = calculate_progress(l3_current, l3_month, l3_target)

        # 计算总进度
        total_progress = calculate_total_progress(l1_progress, l2_progress, l3_progress)

        # 构建列名（如 "2月"、"2月.1"、"2月.2"）
        month_col = month_name
        month_col_l2 = f"{month_name}.1"
        month_col_l3 = f"{month_name}.2"

        month_data.append({
            '系统': system_name,
            '车厢': carriage,
            '产品': product,
            'L1 现状': l1_current,
            month_col: l1_month,
            'L1 目标': l1_target,
            'L1 进度': l1_progress,
            'L2 现状': l2_current,
            month_col_l2: l2_month,
            'L2 目标': l2_target,
            'L2 进度': l2_progress,
            'L3 现状': l3_current,
            month_col_l3: l3_month,
            'L3 目标': l3_target,
            'L3 进度': l3_progress,
            '总进度': total_progress
        })

    # 创建DataFrame
    df_month = pd.DataFrame(month_data)

    # 打开Excel文件
    wb = load_workbook(value_file)

    # 如果sheet已存在，删除旧sheet
    if month_name in wb.sheetnames:
        print(f"\n删除旧的 '{month_name}' sheet")
        del wb[month_name]
        wb.save(value_file)

    # 写入新sheet
    with pd.ExcelWriter(value_file, engine='openpyxl', mode='a') as writer:
        df_month.to_excel(writer, sheet_name=month_name, index=False)

    # 格式化进度列为百分比
    wb = load_workbook(value_file)
    ws = wb[month_name]

    # 找到进度列
    headers = [cell.value for cell in ws[1]]
    progress_cols = [i for i, h in enumerate(headers) if '进度' in str(h)]

    for col_idx in progress_cols:
        col_letter = chr(65 + col_idx)
        for row in range(2, ws.max_row + 1):
            cell = ws[f'{col_letter}{row}']
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    wb.save(value_file)

    print(f"\n✅ 成功创建 '{month_name}' sheet！")
    print()
    print("进度汇总:")
    print("-" * 80)

    for _, row in df_month.iterrows():
        l1_p = row['L1 进度']
        l2_p = row['L2 进度']
        l3_p = row['L3 进度']
        total_p = row['总进度']

        l1_str = f"{l1_p*100:.2f}%" if pd.notna(l1_p) else "N/A"
        l2_str = f"{l2_p*100:.2f}%" if pd.notna(l2_p) else "N/A"
        l3_str = f"{l3_p*100:.2f}%" if pd.notna(l3_p) else "N/A"
        total_str = f"{total_p*100:.2f}%" if pd.notna(total_p) else "N/A"

        print(f"{row['系统']}: L1={l1_str}, L2={l2_str}, L3={l3_str} | 总进度={total_str}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='创建价值分月度sheet')
    parser.add_argument('--value-file', required=True, help='价值分通晒表格.xlsx 路径')
    parser.add_argument('--stats-file', required=True, help='系统价值分统计.xlsx 路径')
    parser.add_argument('--month', required=True, help='月份名称（如 2月、3月）')
    parser.add_argument('--date', help='统计表日期筛选（如 2026-03-02）')

    args = parser.parse_args()

    create_monthly_sheet(
        value_file=args.value_file,
        stats_file=args.stats_file,
        month_name=args.month,
        stats_date=args.date
    )
